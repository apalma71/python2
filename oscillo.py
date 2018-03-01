import vxi11, ctypes, os, shutil, csv, xlsxwriter, openpyxl, Tkinter, tkMessageBox
from openpyxl import load_workbook
#from Tkinter import *

instr =  vxi11.Instrument("172.17.1.18")
#MessageBox = ctypes.windll.user32.MessageBoxW
print(instr.ask("*IDN?"))

wb = load_workbook(filename="Essai1.xlsx")
ws = wb.worksheets[0]


#FPR8
prf = 0

while (prf < 2):
    prf = prf + 1
    voien = -1
    while ( voien < 7):  #Boucle sur 8 voies
        voien = voien + 1
        instr.open()
        ad = 0
        ad2 = 0
        moy1 = 0
        td = 0
        tdm = 0
        tkMessageBox.showinfo(None, "Passer Ã  la voie "+ str (voien))
        #MessageBox(None, 'Passer Ã  la voie ' + str (voien) , 'Voie NÂ°', 0)

        # Test Moyenne pour Tension Ã  50 Ohms

        while(ad < 20):
            instr.open()
            ad = ad + 1
            instr.write("MEASUREMENT:IMMED:TYPE LOW")
            moy = float (instr.ask("MEASUREMENT:IMMED:VALUE?"))
            #print(moy1)
            moy1 = moy1 + moy
            instr.close()
        print ("Tension 50Ohms : " + str((moy1/20)*(-1)) + " V")
        valeur1 = str((moy1/20)*(-1))
        ad = 0
        moy1 = 0

          # Test Moyenne pour Tension Ã  75 Ohms

        #MessageBox(None, 'Passer Ã  75 Ohms ', '50 / 75 Ohms', 0)
        tkMessageBox.showinfo("Passer Ã  75 Ohms", "50 / 75 Ohms" )
        while(ad < 20):
             instr.open()
             ad = ad + 1
             instr.write("MEASUREMENT:IMMED:TYPE LOW")
             moy = float (instr.ask("MEASUREMENT:IMMED:VALUE?"))
             #print(moy1)
             moy1 = moy1 + moy
             instr.close()
        print ("Tension 75Ohms : " + str((moy1/20)*(-1)) + " V")
        valeur2 = str((moy1/20)*(-1))
        tkMessageBox.showinfo("Passer Ã  50 Ohms", "50 / 75 Ohms" )
        #MessageBox(None, 'Passer Ã  50 Ohms ', '50 / 75 Ohms', 0)


        # Test FALL(Temps de descente)
        while(ad2 < 20):
            instr.open()
            ad2 = ad2 + 1
            instr.write("MEASUREMENT:IMMED:TYPE FALL")
            td = float (instr.ask("MEASUREMENT:IMMED:VALUE?"))
            td = (td * 10E8)
            tdm = tdm + td
        valeur3 = str(tdm/20)
        print ("Temps de descente : "+ valeur3)


        # Test DurÃ©e d'impulsion

        instr.write("MEASUREMENT:IMMED:TYPE NWIdth")
        valeur4 = float (instr.ask("MEASUREMENT:IMMED:VALUE?"))
        valeur4 = str(valeur4 * 10E8)
        print("DurÃ©e = " + valeur4)
        instr.close()

        # Export Excel

        if prf == 1:
            if voien == 0:
                    ws['E73'] = float(valeur1)
                    ws['F73'] = float(valeur2)
                    ws['E74'] = float(valeur4)
                    ws['E75'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 1:
                    ws['G73'] = float(valeur1)
                    ws['H73'] = float(valeur2)
                    ws['G74'] = float(valeur4)
                    ws['G75'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 2:
                    ws['I73'] = float(valeur1)
                    ws['J73'] = float(valeur2)
                    ws['I74'] = float(valeur4)
                    ws['I75'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 3:
                    ws['K73'] = float(valeur1)
                    ws['L73'] = float(valeur2)
                    ws['K74'] = float(valeur4)
                    ws['K75'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 4:
                    ws['E80'] = float(valeur1)
                    ws['F80'] = float(valeur2)
                    ws['E81'] = float(valeur4)
                    ws['E82'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 5:
                    ws['G80'] = float(valeur1)
                    ws['H80'] = float(valeur2)
                    ws['G81'] = float(valeur4)
                    ws['G82'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 6:
                    ws['I80'] = float(valeur1)
                    ws['J80'] = float(valeur2)
                    ws['I81'] = float(valeur4)
                    ws['I82'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 7:
                    ws['K80'] = float(valeur1)
                    ws['L80'] = float(valeur2)
                    ws['K81'] = float(valeur4)
                    ws['K82'] = float(valeur3)
                    wb.save("Essai1.xlsx")
        if prf == 2:
            if voien == 0:
                    ws['E90'] = float(valeur1)
                    ws['F90'] = float(valeur2)
                    ws['E91'] = float(valeur4)
                    ws['E92'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 1:
                    ws['G90'] = float(valeur1)
                    ws['H90'] = float(valeur2)
                    ws['G91'] = float(valeur4)
                    ws['G92'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 2:
                    ws['I90'] = float(valeur1)
                    ws['J90'] = float(valeur2)
                    ws['I91'] = float(valeur4)
                    ws['I92'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 3:
                    ws['K90'] = float(valeur1)
                    ws['L90'] = float(valeur2)
                    ws['K91'] = float(valeur4)
                    ws['K92'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 4:
                    ws['E97'] = float(valeur1)
                    ws['F97'] = float(valeur2)
                    ws['E98'] = float(valeur4)
                    ws['E99'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 5:
                    ws['G97'] = float(valeur1)
                    ws['H97'] = float(valeur2)
                    ws['G98'] = float(valeur4)
                    ws['G99'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 6:
                    ws['I97'] = float(valeur1)
                    ws['J97'] = float(valeur2)
                    ws['I98'] = float(valeur4)
                    ws['I99'] = float(valeur3)
                    wb.save("Essai1.xlsx")
            if voien == 7:
                    ws['K97'] = float(valeur1)
                    ws['L97'] = float(valeur2)
                    ws['K98'] = float(valeur4)
                    ws['K99'] = float(valeur3)
                    wb.save("Essai1.xlsx")
        #voien = voien + 1 # compteur boucle sur 8 voies (0 Ã  7)
    if prf == 1:
        tkMessageBox.showinfo("Changer la PRF", "PRF")
        #MessageBox(None, 'Changer la PRF','PRF', 0)
    else:
        tkMessageBox.showinfo("Fin", "Fin")
        #MessageBox(None, 'Fin','PRF', 0)
